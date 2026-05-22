from pathlib import Path
from datetime import datetime
import tempfile
import os
import time

from PIL import Image, ExifTags, UnidentifiedImageError
import pillow_heif
from openpyxl import load_workbook, Workbook

pillow_heif.register_heif_opener()

# =========================
# PATHS
# =========================
BASE_DIR = Path(r"C:\Users\agonz\OneDrive\Documentos\White Door Closet")
RAW_DIR = BASE_DIR / "raw_photos"
JPG_DIR = BASE_DIR / "jpg_photos"
EXCEL_FILE = BASE_DIR / "inventory.xlsx"
SHEET_NAME = "photo_archive"

# =========================
# SETTINGS
# =========================
JPEG_QUALITY = 97

HEADERS = [
    "original_name",
    "converted_name",
    "datetime",
    "datetime_source",
    "reviewed",
    "temp_group",
    "notes",
]

ALLOWED_EXTENSIONS = {".heic", ".heif", ".jpg", ".jpeg"}
EXIF_TAGS = {v: k for k, v in ExifTags.TAGS.items()}


def parse_exif_datetime(dt_str):
    if not dt_str:
        return None
    try:
        return datetime.strptime(dt_str, "%Y:%m:%d %H:%M:%S")
    except Exception:
        return None


def parse_subsec(subsec_value):
    if subsec_value is None:
        return 0

    s = str(subsec_value).strip()
    digits = "".join(ch for ch in s if ch.isdigit())

    if not digits:
        return 0

    if len(digits) >= 3:
        return int(digits[:3])
    elif len(digits) == 2:
        return int(digits) * 10
    else:
        return int(digits) * 100


def is_supported_image(file_path):
    if not file_path.is_file():
        return False

    suffix = file_path.suffix.lower()

    if suffix in ALLOWED_EXTENSIONS:
        return True

    try:
        with Image.open(file_path) as img:
            img.verify()
        return True
    except Exception:
        return False


def get_original_capture_datetime(file_path):
    try:
        with Image.open(file_path) as img:
            exif = img.getexif()

            dt_original = exif.get(EXIF_TAGS.get("DateTimeOriginal"))
            dt_digitized = exif.get(EXIF_TAGS.get("DateTimeDigitized"))
            dt_general = exif.get(EXIF_TAGS.get("DateTime"))

            subsec_original = exif.get(EXIF_TAGS.get("SubsecTimeOriginal"))
            if subsec_original is None:
                subsec_original = exif.get(37521)

            dt = (
                parse_exif_datetime(dt_original)
                or parse_exif_datetime(dt_digitized)
                or parse_exif_datetime(dt_general)
            )

            if dt is not None:
                ms = parse_subsec(subsec_original)
                return dt, ms, "metadata"

    except Exception:
        pass

    dt = datetime.fromtimestamp(file_path.stat().st_mtime)
    ms = int(dt.microsecond / 1000)
    return dt, ms, "filesystem_fallback"


def build_output_name(dt, ms):
    return f"WDC_{dt.strftime('%y%m%d_%H%M%S')}_{ms:03d}.jpg"


def convert_image_to_jpg(src, dest):
    with Image.open(src) as img:
        rgb = img.convert("RGB")
        rgb.save(dest, "JPEG", quality=JPEG_QUALITY, subsampling=0)


def open_or_create_excel():
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.create_sheet(SHEET_NAME)

    ensure_headers(ws)
    return wb, ws


def ensure_headers(ws):
    for col, header in enumerate(HEADERS, start=1):
        current = ws.cell(row=1, column=col).value
        if current is None or str(current).strip() == "":
            ws.cell(row=1, column=col, value=header)


def get_header_map(ws):
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value).strip()] = cell.column
    return header_map


def get_existing_original_names(ws):
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        original_name = row[0]
        if original_name:
            existing.add(str(original_name).strip())
    return existing


def get_existing_converted_names(ws):
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        converted_name = row[1]
        if converted_name:
            existing.add(str(converted_name).strip())
    return existing


def safe_save_workbook(wb, target_file):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        temp_name = tmp.name

    try:
        wb.save(temp_name)
        os.replace(temp_name, target_file)
    finally:
        if os.path.exists(temp_name):
            try:
                os.remove(temp_name)
            except Exception:
                pass


def repair_missing_jpgs(ws):
    """
    If photo_archive already has:
        original_name + converted_name
    but jpg_photos/converted_name is missing
    and raw_photos/original_name exists,
    recreate the JPG WITHOUT adding a new Excel row.
    """
    header_map = get_header_map(ws)

    original_col = header_map.get("original_name")
    converted_col = header_map.get("converted_name")
    notes_col = header_map.get("notes")
    datetime_source_col = header_map.get("datetime_source")

    if original_col is None or converted_col is None:
        raise ValueError("photo_archive must have original_name and converted_name columns.")

    repaired = 0
    cannot_repair = []

    for row_idx in range(2, ws.max_row + 1):
        original_name = ws.cell(row=row_idx, column=original_col).value
        converted_name = ws.cell(row=row_idx, column=converted_col).value

        if not original_name or not converted_name:
            continue

        original_name = str(original_name).strip()
        converted_name = str(converted_name).strip()

        raw_path = RAW_DIR / original_name
        jpg_path = JPG_DIR / converted_name

        if jpg_path.exists():
            continue

        if not raw_path.exists():
            cannot_repair.append((original_name, converted_name, "raw file missing"))
            continue

        try:
            convert_image_to_jpg(raw_path, jpg_path)
            repaired += 1

            if notes_col is not None:
                old_note = ws.cell(row=row_idx, column=notes_col).value
                repair_note = "repaired missing jpg"
                if old_note:
                    ws.cell(row=row_idx, column=notes_col, value=f"{old_note}; {repair_note}")
                else:
                    ws.cell(row=row_idx, column=notes_col, value=repair_note)

            if datetime_source_col is not None:
                old_source = ws.cell(row=row_idx, column=datetime_source_col).value
                if not old_source:
                    ws.cell(row=row_idx, column=datetime_source_col, value="repaired_existing_row")

        except Exception as e:
            cannot_repair.append((original_name, converted_name, str(e)))

    return repaired, cannot_repair


def main():
    start = time.perf_counter()

    JPG_DIR.mkdir(exist_ok=True)

    wb, ws = open_or_create_excel()

    # 1. Repair missing JPGs from existing photo_archive rows
    repaired_jpgs, cannot_repair = repair_missing_jpgs(ws)

    # 2. Refresh existing sets AFTER repair
    existing_original_names = get_existing_original_names(ws)
    existing_converted_names = get_existing_converted_names(ws)

    all_items = list(RAW_DIR.iterdir())
    files = [p for p in all_items if is_supported_image(p)]

    if not files:
        safe_save_workbook(wb, EXCEL_FILE)
        print("No supported image files found.")
        print(f"Repaired missing JPGs: {repaired_jpgs}")
        return

    # 3. Only process truly new raw files
    new_files = [p for p in files if p.name not in existing_original_names]
    skipped_count = len(files) - len(new_files)

    files_with_dt = []

    for file in new_files:
        dt, ms, source = get_original_capture_datetime(file)
        files_with_dt.append((file, dt, ms, source))

    files_with_dt.sort(key=lambda x: (x[1], x[2], x[0].name.lower()))

    jpgs_created = 0
    rows_added = 0
    skipped_existing_output = 0
    unsupported_or_failed = 0
    collision_files = []

    for file, dt, ms, source in files_with_dt:
        new_name = build_output_name(dt, ms)
        output_path = JPG_DIR / new_name

        # If output name already exists, do NOT add duplicate row.
        # This protects photo_archive from accidental duplicates.
        if output_path.exists() or new_name in existing_converted_names:
            skipped_existing_output += 1
            collision_files.append((file.name, new_name))
            continue

        try:
            convert_image_to_jpg(file, output_path)
        except (UnidentifiedImageError, OSError, ValueError):
            unsupported_or_failed += 1
            continue

        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=file.name)
        ws.cell(row=row, column=2, value=output_path.name)
        ws.cell(row=row, column=3, value=dt.strftime("%Y-%m-%d %H:%M:%S") + f".{ms:03d}")
        ws.cell(row=row, column=4, value=source)
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=7, value="")

        existing_original_names.add(file.name)
        existing_converted_names.add(output_path.name)

        jpgs_created += 1
        rows_added += 1

    safe_save_workbook(wb, EXCEL_FILE)

    elapsed = time.perf_counter() - start

    print(f"Supported image files found: {len(files)}")
    print(f"Already archived raw files skipped: {skipped_count}")
    print(f"Missing JPGs repaired from existing rows: {repaired_jpgs}")
    print(f"New JPGs created: {jpgs_created}")
    print(f"New rows added to photo_archive: {rows_added}")
    print(f"Skipped because output JPG already exists: {skipped_existing_output}")
    print(f"Unsupported/failed during conversion: {unsupported_or_failed}")
    print(f"Elapsed time: {elapsed:.2f} seconds")

    if cannot_repair:
        print("\nCould not repair these missing JPGs:")
        for original_name, converted_name, reason in cannot_repair:
            print(f"- {original_name} -> {converted_name}: {reason}")

    if collision_files:
        print("\nNew raw files with output-name collisions; no rows added:")
        for original_name, target_name in collision_files:
            print(f"- {original_name} -> {target_name}")


if __name__ == "__main__":
    main()